
import openpyxl, pprint
print('Opening workbook...')
wb = openpyxl.load_workbook('Notlar.xlsx')
sheet = wb.get_sheet_by_name('BIL2001')

studentData = []
snfort=0
for row in range(1, sheet.max_row + 1):
 snfort+=(float(sheet['B' + str(row)].value)/194 + float(sheet['C' + str(row)].value)/194)
final=sheet['C' + str(row)].value
if final > 45 :
      print((float(sheet['B' + str(row)].value)/194 + float(sheet['C' + str(row)].value)/194))
print("snfort",snfort)
print("final",final)
stsapma=21.23

print('Reading rows...')
for row in range(1, sheet.max_row + 1):  # sheet.get_highest_row()

    numara = sheet['A' + str(row)].value
    vize   = sheet['B' + str(row)].value
    final  = sheet['C' + str(row)].value
    ortalama= sheet['D' + str(row)].value  = (float(sheet['B' + str(row)].value)/2 + float(sheet['C' + str(row)].value)/2)
    Tnotu= sheet['F' + str(row)].value  = ((int(sheet['B' + str(row)].value)/2 + int(sheet['C' + str(row)].value)/2)- float(snfort))/float(stsapma)*10 +50
    if Tnotu < 45 :
         print(str(numara) +' vize:  '+ str(vize) +' final:  '+ str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'    +str(0)+'harfnotu:'+str('FF')) 

         item = [numara, vize, final,ortalama,Tnotu,'FF']
         studentData.insert(row, item)

    harfnotu= sheet['E' + str(row)].value



    if final <45 :
      print(str(numara) +' vize:  '+ str(vize) +' final:  '+ str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(0)+'harfnotu:'+str('FF')) 

      item = [numara, vize, final,ortalama,Tnotu,'FF']
      studentData.insert(row, item)


    elif final >=45  and Tnotu >=41 and Tnotu<46.99:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final) +' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('DC'))
                                                                
      item = [numara, vize, final,ortalama,Tnotu,'DC']
      studentData.insert(row, item)

    elif   final >=45  and Tnotu >=47 and Tnotu<51.99:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('CC'))

      item = [numara, vize, final,ortalama,Tnotu,'CC']
      studentData.insert(row, item)                                                                                                               

    elif final >=45  and Tnotu >=52 and Tnotu<56.99:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('CB'))

      item = [numara, vize, final,ortalama,Tnotu,'CB']
      studentData.insert(row, item)

    elif   final >=45  and Tnotu >=57 and Tnotu<61.99:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('BB'))

      item = [numara, vize, final,ortalama,Tnotu,'BB']
      studentData.insert(row, item)


    elif   final >=45  and Tnotu >=62 and Tnotu<66.99:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('BA'))

      item = [numara, vize, final,ortalama,Tnotu,'BA']
      studentData.insert(row, item)
 
    elif final>=45 and Tnotu>=67:

      print (str(numara) + 'vize:'+  str(vize) + 'final:'  + str(final)+' ortalama:'+ str(ortalama)+'Tnotu:'+str(Tnotu)+'harfnotu:'+str('AA'))

      item = [numara, vize, final,ortalama,Tnotu,'AA']
      studentData.insert(row, item)



          
print('Writing results...')

max_row = sheet.max_row # Notlar.xlsx

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

for row in range(1, max_row + 1):
    sheet['A' + str(row)] = studentData[row-1][0]
    sheet['B' + str(row)] = studentData[row-1][1]
    sheet['C' + str(row)] = studentData[row-1][2]
    sheet['D' + str(row)] = studentData[row-1][3]
    sheet['E' + str(row)] = studentData[row-1][4]
    sheet['F' + str(row)] = studentData[row-1][5]
    
wb.save('Notlar_4.xlsx')

print('Done.')
print("YÜSEK BÝLGÝSAYAR MÜHENDÝSÝ ENES KARALÝ")
